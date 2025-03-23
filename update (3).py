import tkinter as tk
from tkinter import ttk, messagebox
import sqlite3
from openpyxl import load_workbook
import random
from openpyxl.styles import Alignment
import os


class LoginWindow:
    def __init__(self, root):
        self.root = root
        self.root.title("Admin Girişi")
        self.root.geometry("300x200")
        
        # Veritabanı bağlantısı
        self.conn = sqlite3.connect('universite.db')
        self.cursor = self.conn.cursor()
        
        # GUI bileşenlerini oluştur
        self.init_ui()
        
    def init_ui(self):
        # Ana Frame
        self.main_frame = ttk.Frame(self.root, padding=20)
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Kullanıcı Adı
        ttk.Label(self.main_frame, text="Kullanıcı Adı:").pack(pady=5)
        self.username_var = tk.StringVar()
        self.username_entry = ttk.Entry(self.main_frame, textvariable=self.username_var)
        self.username_entry.pack(pady=5)
        
        # Şifre
        ttk.Label(self.main_frame, text="Şifre:").pack(pady=5)
        self.password_var = tk.StringVar()
        self.password_entry = ttk.Entry(self.main_frame, textvariable=self.password_var, show="*")
        self.password_entry.pack(pady=5)
        
        # Giriş Butonu
        ttk.Button(self.main_frame, text="Giriş Yap", command=self.login).pack(pady=20)
        
    def login(self):
        username = self.username_var.get()
        password = self.password_var.get()
        
        # Veritabanında kullanıcıyı kontrol et
        self.cursor.execute("SELECT * FROM admin WHERE kullanici_adi = ? AND sifre = ?", (username, password))
        user = self.cursor.fetchone()
        
        if user:
            self.root.destroy()  # Login penceresini kapat
            root = tk.Tk()  # Yeni pencere oluştur
            app = DersSecimUygulamasi(root)  # Ana uygulamayı başlat
            root.mainloop()
        else:
            messagebox.showerror("Hata", "Kullanıcı adı veya şifre hatalı!")
    
    def __del__(self):
        self.conn.close()


class DersSecimUygulamasi:
    def __init__(self, root):
        self.root = root
        self.root.title("Ders Seçim ve Program Oluşturma Sistemi")
        self.root.geometry("1200x800")

        # Veritabanı bağlantıları
        self.conn = sqlite3.connect('universite.db')
        self.cursor = self.conn.cursor()

        # Derslik verilerini yükle
        self.derslikler = self.derslikleri_yukle()

        # GUI bileşenlerini oluştur
        self.init_ui()

    def init_ui(self):
        # Ana Frame
        self.main_frame = ttk.Frame(self.root, padding=20)
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        # Bölüm Seçimi
        ttk.Label(self.main_frame, text="Bölüm:").grid(row=0, column=0, sticky=tk.W)
        self.bolum_var = tk.StringVar()
        self.bolum_combo = ttk.Combobox(self.main_frame, textvariable=self.bolum_var,
                                        values=['Yazılım Mühendisliği', 'Bilgisayar Mühendisliği'])
        self.bolum_combo.grid(row=0, column=1, sticky=tk.EW)

        # Dönem Seçimi
        ttk.Label(self.main_frame, text="Dönem:").grid(row=1, column=0, sticky=tk.W)
        self.donem_var = tk.StringVar()
        self.donem_combo = ttk.Combobox(self.main_frame, textvariable=self.donem_var)
        self.donem_combo.grid(row=1, column=1, sticky=tk.EW)
        self.bolum_combo.bind('<<ComboboxSelected>>', self.donemleri_guncelle)

        # Öğrenci Sayısı
        ttk.Label(self.main_frame, text="Öğrenci Sayısı:").grid(row=2, column=0, sticky=tk.W)
        self.ogrenci_sayisi = ttk.Entry(self.main_frame)
        self.ogrenci_sayisi.grid(row=2, column=1, sticky=tk.EW)

        # Ders Listesi
        self.tree = ttk.Treeview(self.main_frame, columns=('Kod', 'Ad', 'AKTS', 'Tip', 'Saat', 'Hoca', 'ders_sekli'), show='headings')
        self.tree.heading('Kod', text='Ders Kodu')
        self.tree.heading('Ad', text='Ders Adı')
        self.tree.heading('AKTS', text='AKTS')
        self.tree.heading('Tip', text='Tip')
        self.tree.heading('Saat', text='Saat')
        self.tree.heading('Hoca', text='Öğretim Üyesi')
        self.tree.heading('ders_sekli', text='Öğrenim')
        self.tree.grid(row=3, column=0, columnspan=2, pady=10, sticky=tk.NSEW)

        # Excel'e Aktar Butonu
        ttk.Button(self.main_frame, text="Excel'e Aktar", command=self.excel_aktar).grid(row=4, column=0, columnspan=2)

        # Grid konfigürasyonu
        self.main_frame.columnconfigure(1, weight=1)
        self.main_frame.rowconfigure(3, weight=1)

    def derslikleri_yukle(self):
        try:
            self.cursor.execute("SELECT derslik_id, kapasite, statü FROM derslikler")
            return self.cursor.fetchall()
        except sqlite3.OperationalError as e:
            messagebox.showerror("Veritabanı Hatası", f"Tablo veya sütun bulunamadı: {str(e)}")
            return []

    def uygun_derslik_bul(self, ogrenci_sayisi):
        for derslik in sorted(self.derslikler, key=lambda x: x[1]):
            if derslik[1] >= ogrenci_sayisi:
                return derslik
        return None

    def donemleri_guncelle(self, event):
        self.donem_combo['values'] = list(range(1, 9))
        self.donem_combo.bind('<<ComboboxSelected>>', self.dersleri_goster)

    def dersleri_goster(self, event):
        self.tree.delete(*self.tree.get_children())
        bolum = self.bolum_var.get()
        donem = self.donem_var.get()

        # Bölüme özel dersleri yükle
        tablo = 'yazilim_muh_dersler' if bolum == 'Yazılım Mühendisliği' else 'bilgisayar_muh_dersler'
        self.cursor.execute(f'''
            SELECT ders_kodu, ders_adi, akts, ders_tipi, saat, ders_hocasi, ders_sekli 
            FROM {tablo} 
            WHERE donem = ?''', (donem,))

        for ders in self.cursor.fetchall():
            self.tree.insert('', tk.END, values=ders + ('',))

        # Ortak dersleri yükle
        self.cursor.execute('''
            SELECT ders_kodu, ders_adi, akts, ders_tipi, saat ,ders_hocasi, ders_sekli
            FROM ortak_dersler 
            WHERE donem = ?''', (donem,))

        for ders in self.cursor.fetchall():
            self.tree.insert('', tk.END, values=ders + ('',))

    def excel_aktar(self):
        try:
            ogrenci_sayisi = int(self.ogrenci_sayisi.get())
            secili_derslik = self.uygun_derslik_bul(ogrenci_sayisi)

            if not secili_derslik:
                messagebox.showerror("Hata", "Uygun derslik bulunamadı!")
                return

            secili_itemler = self.tree.selection()
            if not secili_itemler:
                messagebox.showerror("Hata", "Lütfen ders seçiniz!")
                return

            # Bölüm adını al ve dosya adını oluştur
            bolum = self.bolum_var.get()
            excel_dosya_adi = f"{bolum}_Ders_Programı.xlsx"

            # Excel dosyasını yükle
            if os.path.exists(excel_dosya_adi):
                wb = load_workbook(excel_dosya_adi)
            else:
                wb = load_workbook('Ders_Programı.xlsx')
            ws = wb.active

            gunler = ['Pazartesi', 'Salı', 'Çarşamba', 'Perşembe', 'Cuma']
            gun_satirlari = {'Pazartesi': 4, 'Salı': 16, 'Çarşamba': 28, 'Perşembe': 40, 'Cuma': 52}
            uzaktan_satirlari = {'Pazartesi': 12, 'Salı': 24, 'Çarşamba': 36, 'Perşembe': 48, 'Cuma': 60}
            musait_saatler = {gun: [True] * 10 for gun in gunler}
            uzaktan_musait_saatler = {gun: [True] * 4 for gun in gunler}

            # Önceki dersleri kontrol et ve musait saatleri güncelle
            for gun in gunler:
                # Normal dersler için kontrol
                baslangic_satiri = gun_satirlari[gun]
                for i in range(10):
                    satir = baslangic_satiri + i
                    hucre = f'{self.sinif_sec()}{satir}'
                    if ws[hucre].value:
                        musait_saatler[gun][i] = False

                # Uzaktan dersler için kontrol
                baslangic_satiri = uzaktan_satirlari[gun]
                for i in range(4):
                    satir = baslangic_satiri + i
                    hucre = f'{self.sinif_sec()}{satir}'
                    if ws[hucre].value:
                        uzaktan_musait_saatler[gun][i] = False

            # Yeni dersleri ekle
            for item in secili_itemler:
                ders = self.tree.item(item)['values']
                ders_saati = int(ders[4])
                ders_sekli = ders[6]

                if ders_sekli == "uzaktan":
                    # Uzaktan dersler için özel zamanlama
                    for gun in random.sample(gunler, len(gunler)):
                        for i in range(len(uzaktan_musait_saatler[gun]) - ders_saati + 1):
                            if all(uzaktan_musait_saatler[gun][i:i + ders_saati]):
                                satir = uzaktan_satirlari[gun] + i
                                hucre = f'{self.sinif_sec()}{satir}'

                                ws[hucre] = f"{ders[0]} - {ders[1]}\n{ders[5]}\nUzaktan"
                                ws[hucre].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                                ws.merge_cells(f"{hucre}:{self.sinif_sec()}{satir + ders_saati - 1}")

                                for j in range(i, i + ders_saati):
                                    uzaktan_musait_saatler[gun][j] = False
                                break
                        else:
                            continue
                        break
                else:
                    # Normal dersler için normal zamanlama
                    for gun in random.sample(gunler, len(gunler)):
                        for i in range(len(musait_saatler[gun]) - ders_saati + 1):
                            if all(musait_saatler[gun][i:i + ders_saati]):
                                satir = gun_satirlari[gun] + i
                                hucre = f'{self.sinif_sec()}{satir}'

                                ws[hucre] = f"{ders[0]} - {ders[1]}\n{ders[5]}\nDerslik: {secili_derslik[0]}"
                                ws[hucre].alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                                ws.merge_cells(f"{hucre}:{self.sinif_sec()}{satir + ders_saati - 1}")

                                for j in range(i, i + ders_saati):
                                    musait_saatler[gun][j] = False
                                break
                        else:
                            continue
                        break

            # Dosyayı kaydet
            wb.save(excel_dosya_adi)

            # Seçili dersleri Treeview'dan kaldır
            for item in secili_itemler:
                self.tree.delete(item)

            messagebox.showinfo("Başarılı", f"Dersler {excel_dosya_adi} dosyasına aktarıldı ve listeden kaldırıldı!")

        except ValueError:
            messagebox.showerror("Hata", "Geçersiz öğrenci sayısı!")
        except Exception as e:
            messagebox.showerror("Hata", f"Bir hata oluştu: {str(e)}")

    def sinif_sec(self):
        donem = int(self.donem_var.get())
        return 'C' if donem <= 2 else 'D' if donem <= 4 else 'E' if donem <= 6 else 'F'

    def _del_(self):
        self.conn.close()


if __name__ == "__main__":
    root = tk.Tk()
    login = LoginWindow(root)
    root.mainloop()